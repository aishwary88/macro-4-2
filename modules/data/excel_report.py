"""
Excel report generation using OpenPyXL.
Creates formatted Excel files with vehicle data, conditional formatting,
and summary analytics.
"""

from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter
from modules.utils.logger import get_logger
from core.constants import EXCEL_RED, EXCEL_GREEN, EXCEL_HEADER_BG, EXCEL_HEADER_FG

logger = get_logger("excel_report")

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color=EXCEL_HEADER_FG)
HEADER_FILL = PatternFill(start_color=EXCEL_HEADER_BG, end_color=EXCEL_HEADER_BG, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

RED_FILL = PatternFill(start_color=EXCEL_RED, end_color=EXCEL_RED, fill_type="solid")
GREEN_FILL = PatternFill(start_color=EXCEL_GREEN, end_color=EXCEL_GREEN, fill_type="solid")


def generate_excel_report(
    vehicle_df: pd.DataFrame,
    analytics: dict,
    output_path: str,
    video_filename: str = "Unknown",
) -> str:
    """Generate a formatted Excel report.

    Args:
        vehicle_df: DataFrame with vehicle data.
        analytics: Dict with analytics data.
        output_path: Path to save the Excel file.
        video_filename: Name of the source video.

    Returns:
        Path to the generated Excel file.
    """
    wb = Workbook()

    # ---- Sheet 1: Vehicle Log ----
    ws_vehicles = wb.active
    ws_vehicles.title = "Vehicle Log"
    _create_vehicle_sheet(ws_vehicles, vehicle_df)

    # ---- Sheet 2: Summary ----
    ws_summary = wb.create_sheet("Summary")
    _create_summary_sheet(ws_summary, analytics, video_filename)

    # ---- Sheet 3: Speed Details ----
    if not vehicle_df.empty:
        ws_speed = wb.create_sheet("Speed Details")
        _create_speed_details_sheet(ws_speed, vehicle_df)

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Excel report saved: {output_path}")
    return output_path


def _create_vehicle_sheet(ws, df: pd.DataFrame):
    """Create the main vehicle log sheet."""
    headers = [
        "Vehicle ID", "Type", "Plate Number", "Avg Speed (km/h)",
        "Max Speed (km/h)", "Status", "First Seen", "Last Seen", "Frames"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_overspeed = row.get("overspeed", False)
        row_fill = RED_FILL if is_overspeed else GREEN_FILL
        status_text = "🔴 OVERSPEED" if is_overspeed else "🟢 NORMAL"

        data = [
            row.get("vehicle_id", ""),
            row.get("vehicle_type", ""),
            row.get("plate_number", "N/A"),
            row.get("avg_speed", 0),
            row.get("max_speed", 0),
            status_text,
            row.get("first_seen", ""),
            row.get("last_seen", ""),
            row.get("frame_count", 0),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

            # Apply row fill for status column and speed columns
            if col in [4, 5, 6]:
                cell.fill = row_fill
                if is_overspeed:
                    cell.font = Font(bold=True, color="FFFFFF")

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_width = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_width + 4, 30)


def _create_summary_sheet(ws, analytics: dict, video_filename: str):
    """Create the summary analytics sheet."""
    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="📊 Traffic Analysis Summary")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1E293B")
    title_cell.alignment = Alignment(horizontal="center")

    # Video info
    ws.cell(row=3, column=1, value="Video:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=video_filename)

    # Stats
    stats = [
        ("Total Vehicles", analytics.get("total_vehicles", 0)),
        ("Overspeed Count", analytics.get("overspeed_count", 0)),
        ("Average Speed (km/h)", analytics.get("avg_speed", 0)),
        ("Max Speed (km/h)", analytics.get("max_speed", 0)),
        ("Min Speed (km/h)", analytics.get("min_speed", 0)),
    ]

    row = 5
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # Vehicle type breakdown
    row += 1
    ws.cell(row=row, column=1, value="Vehicle Types").font = Font(bold=True, size=14)
    row += 1

    for vtype, count in analytics.get("vehicle_types", {}).items():
        ws.cell(row=row, column=1, value=vtype).font = Font(bold=True)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


def _create_speed_details_sheet(ws, df: pd.DataFrame):
    """Create speed details sheet with per-vehicle speed data."""
    headers = ["Vehicle ID", "Type", "Avg Speed", "Max Speed", "Overspeed"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row.get("vehicle_id", ""))
        ws.cell(row=row_idx, column=2, value=row.get("vehicle_type", ""))
        ws.cell(row=row_idx, column=3, value=round(row.get("avg_speed", 0), 2))
        ws.cell(row=row_idx, column=4, value=round(row.get("max_speed", 0), 2))

        overspeed_cell = ws.cell(
            row=row_idx, column=5,
            value="YES" if row.get("overspeed", False) else "NO"
        )
        overspeed_cell.fill = RED_FILL if row.get("overspeed", False) else GREEN_FILL
